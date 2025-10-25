"""
XLSX translation file handler for multi-language translations
"""

import os
from typing import Dict, Any, List, Optional
from pathlib import Path
from openpyxl import load_workbook, Workbook


def read_xlsx_file(file_path: str) -> Dict[str, Any]:
    """
    Read an XLSX translation file and return its content as a dictionary.
    
    Args:
        file_path: Path to the XLSX file
        
    Returns:
        Dictionary containing the XLSX file content with language columns
        
    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file is not valid XLSX format
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"XLSX file not found: {file_path}")
    
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active
        
        if ws is None:
            raise ValueError(f"No active worksheet found in XLSX file {file_path}")
        
        # Get all rows as a list
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        
        if not rows:
            raise ValueError(f"XLSX file {file_path} is empty")
        
        # First row should be headers
        headers = [_safe_cell_value(cell) for cell in rows[0]]
        if len(headers) < 2:
            raise ValueError(f"XLSX file {file_path} must have at least 2 columns (key and at least one language)")
        
        # First column is the key, rest are language codes
        key_column = headers[0]
        language_columns = headers[1:]
        
        # Validate language codes
        for lang in language_columns:
            if not lang or not str(lang).strip():
                raise ValueError(f"Empty language code found in header: {headers}")
        
        # Parse data rows
        translations = {}
        for row in rows[1:]:
            if not row:  # Skip empty rows
                continue
            
            # Convert all cells to safe values
            safe_row = [_safe_cell_value(cell) for cell in row]
            
            # Skip completely empty rows
            if all(not cell for cell in safe_row):
                continue
            
            if len(safe_row) != len(headers):
                continue  # Skip malformed rows
            
            key = safe_row[0].strip() if safe_row[0] else ""
            if not key:
                continue  # Skip rows without keys
            
            # Create language translations
            lang_translations = {}
            for i, lang in enumerate(language_columns):
                value = safe_row[i + 1].strip() if i + 1 < len(safe_row) and safe_row[i + 1] else ""
                if value:  # Only include non-empty translations
                    lang_translations[lang] = value
            
            if lang_translations:  # Only include keys with at least one translation
                translations[key] = lang_translations
        
        return {
            'key_column': key_column,
            'languages': language_columns,
            'translations': translations
        }
    
    except Exception as e:
        raise ValueError(f"Failed to parse XLSX file {file_path}: {str(e)}")


def write_xlsx_file(file_path: str, content: Dict[str, Any]) -> None:
    """
    Write content to an XLSX translation file.
    
    Args:
        file_path: Path where to write the XLSX file
        content: Dictionary containing the XLSX content
        
    Raises:
        ValueError: If content is not a valid dictionary
    """
    if not isinstance(content, dict):
        raise ValueError("XLSX content must be a dictionary")
    
    if 'translations' not in content or 'languages' not in content:
        raise ValueError("XLSX content must contain 'translations' and 'languages' keys")
    
    # Ensure the directory exists
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    key_column = content.get('key_column', 'key')
    languages = content['languages']
    translations = content['translations']
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"
    
    # Write header
    header_row = [key_column] + languages
    ws.append(header_row)
    
    # Write data rows
    for key, lang_translations in translations.items():
        if isinstance(lang_translations, dict):
            row = [key]
            for lang in languages:
                row.append(lang_translations.get(lang, ""))
            ws.append(row)
    
    # Save the workbook
    wb.save(file_path)


def extract_translatable_strings(xlsx_content: Dict[str, Any], 
                               target_language: str) -> Dict[str, str]:
    """
    Extract translatable strings from XLSX content for a specific language.
    
    Args:
        xlsx_content: XLSX file content as dictionary
        target_language: Target language code
        
    Returns:
        Dictionary of key-value pairs for translatable strings
    """
    if 'translations' not in xlsx_content:
        return {}
    
    translatable = {}
    translations = xlsx_content['translations']
    
    for key, lang_translations in translations.items():
        if isinstance(lang_translations, dict) and target_language in lang_translations:
            translatable[key] = lang_translations[target_language]
    
    return translatable


def create_xlsx_from_translations(translations: Dict[str, str], 
                                languages: List[str],
                                key_column: str = "key") -> Dict[str, Any]:
    """
    Create XLSX content from translations.
    
    Args:
        translations: Dictionary of key-value translation pairs
        languages: List of language codes
        key_column: Name of the key column
        
    Returns:
        XLSX content dictionary
    """
    # For single language translations, we need to create a structure
    # that can be extended with more languages
    xlsx_translations = {}
    
    for key, value in translations.items():
        # Create a single-language translation structure
        lang_translations = {languages[0]: value} if languages else {}
        xlsx_translations[key] = lang_translations
    
    return {
        'key_column': key_column,
        'languages': languages,
        'translations': xlsx_translations
    }


def add_language_to_xlsx(xlsx_content: Dict[str, Any], 
                        language: str, 
                        translations: Dict[str, str]) -> Dict[str, Any]:
    """
    Add a new language to existing XLSX content.
    
    Args:
        xlsx_content: Existing XLSX content
        language: New language code to add
        translations: Translations for the new language
        
    Returns:
        Updated XLSX content dictionary
    """
    if 'translations' not in xlsx_content:
        xlsx_content['translations'] = {}
    
    if language not in xlsx_content['languages']:
        xlsx_content['languages'].append(language)
    
    for key, value in translations.items():
        if key not in xlsx_content['translations']:
            xlsx_content['translations'][key] = {}
        xlsx_content['translations'][key][language] = value
    
    return xlsx_content


def get_xlsx_language_codes(xlsx_content: Dict[str, Any]) -> List[str]:
    """
    Get all language codes from XLSX content.
    
    Args:
        xlsx_content: XLSX file content as dictionary
        
    Returns:
        List of language codes
    """
    return xlsx_content.get('languages', [])


def is_valid_xlsx_file(file_path: str) -> bool:
    """
    Check if a file is a valid XLSX translation file.
    
    Args:
        file_path: Path to the file to check
        
    Returns:
        True if the file is a valid XLSX file, False otherwise
    """
    try:
        content = read_xlsx_file(file_path)
        return 'translations' in content and 'languages' in content
    except (FileNotFoundError, ValueError):
        return False


def get_xlsx_language_code(file_path: str) -> Optional[str]:
    """
    Extract language code from XLSX file path.
    
    Args:
        file_path: Path to the XLSX file
        
    Returns:
        Language code if found, None otherwise
    """
    # For XLSX files, we typically don't extract language from filename
    # since they contain multiple languages. Return None to indicate
    # this is a multi-language file.
    return None


def is_glossary_xlsx(file_path: str) -> bool:
    """
    Check if an XLSX file is a glossary file (vs translation file).
    
    Args:
        file_path: Path to the XLSX file
        
    Returns:
        True if the file appears to be a glossary file, False otherwise
    """
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active
        
        if ws is None:
            wb.close()
            return False
        
        # Get first row
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        
        if not rows:
            return False
        
        headers = [_safe_cell_value(cell) for cell in rows[0]]
        
        # Glossary files typically have "Record ID" as first column
        # Translation files typically have "key" as first column
        if headers and len(headers) > 0:
            first_header = str(headers[0]).lower().strip()
            return first_header in ['record id', 'record_id', 'id']
    except Exception:
        pass
    
    return False


def _safe_cell_value(cell_value) -> str:
    """
    Convert a cell value to a safe string representation.
    
    Args:
        cell_value: The cell value to convert
        
    Returns:
        String representation of the cell value
    """
    if cell_value is None:
        return ""
    
    if isinstance(cell_value, bool):
        return "TRUE" if cell_value else "FALSE"
    
    if isinstance(cell_value, (int, float)):
        if isinstance(cell_value, float):
            if cell_value != cell_value:  # NaN check
                return ""
            if cell_value.is_integer():
                return str(int(cell_value))
        return str(cell_value)
    
    if isinstance(cell_value, str):
        return cell_value.strip()
    
    return str(cell_value)
